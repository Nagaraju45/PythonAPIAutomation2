from openpyxl import load_workbook


class Common():
    def __init__(self, file_name, sheet):
        global wb
        global sh
        wb = load_workbook(file_name)
        sh = wb[sheet]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
        return li

    def update_request_with_data(self, row_number, json_request, key_list):
        c = sh.max_column
        for i in range(1, c + 1):
            cells = sh.cell(row=row_number, column=i)
            json_request[key_list[i - 1]] = cells.value
        return json_request

# def read_from_excel_data(self, file_name, sheet):
#     data_list = []
#     wb = load_workbook(file_name)
#     sh = wb[sheet]
#     row_count = sh.max_row
#     col_count = sh.max_column
#     for i in range(2, row_count + 1):
#         row = []
#         for j in range(1, col_count + 1):
#             row.append(sh.cell(row=i, column=j).value)
#         data_list.append(row)
#     return data_list
