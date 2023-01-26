import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
    url = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("C:/APIAutomation2/files/add_new_student.json", "r")
    json_requests = json.loads(f.read())
    wb = openpyxl.load_workbook("C:/Users/nagar/OneDrive/Documents/add_students.xlsx")
    sh = wb['Sheet1']
    rows = sh.max_row
    col = sh.max_column
    for i in range(2, rows + 1):
        # for j in range(1, col + 1):
        #     res = (sh.cell(row=i, column=j).value)
        #     json_requests[col] = res

        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_requests['first_name'] = cell_first_name.value
        json_requests['middle_name'] = cell_middle_name.value
        json_requests['last_name'] = cell_last_name.value
        json_requests['date_of_birth'] = cell_dob.value

        response = requests.post(url, json_requests)
        print(response.text)
        assert response.status_code == 201
